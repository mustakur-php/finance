from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from .models import Client, Activity
from accounts.decorators import admin_required, sales_required


def get_tenant_clients(user, client_type=Client.TYPE_ACTUAL):
    from django.db.models import Q
    qs = Client.objects.filter(tenant=user.tenant, is_active=True, client_type=client_type)
    if user.is_sales:
        qs = qs.filter(assigned_sales=user)
    elif user.is_accountant:
        qs = qs.filter(assigned_accountant=user)
    elif user.is_review and not user.is_admin:
        qs = qs.filter(Q(assigned_sales=user) | Q(created_by=user))
    return qs


def can_access_client(user, client):
    """يتحقق أن للمستخدم صلاحية الوصول لعميل محدد حسب دوره."""
    if user.is_admin:
        return True
    if user.is_sales:
        return client.assigned_sales_id == user.id
    if user.is_accountant:
        return client.assigned_accountant_id == user.id
    if user.is_review:
        return client.assigned_sales_id == user.id or client.created_by_id == user.id
    return False


@login_required
def clients_list(request):
    from accounts.models import User as UserModel
    clients = get_tenant_clients(request.user, Client.TYPE_ACTUAL)
    q = request.GET.get('q', '')
    city = request.GET.get('city', '')
    district = request.GET.get('district', '')
    activity = request.GET.get('activity', '')
    sales_id = request.GET.get('sales', '')
    accountant_id = request.GET.get('accountant', '')
    if q:
        clients = clients.filter(name__icontains=q) | clients.filter(company__icontains=q)
    if city:
        clients = clients.filter(city=city)
    if district:
        clients = clients.filter(district__icontains=district)
    if activity:
        clients = clients.filter(activity_id=activity)
    if sales_id and request.user.is_admin:
        clients = clients.filter(assigned_sales_id=sales_id)
    if accountant_id and request.user.is_admin:
        clients = clients.filter(assigned_accountant_id=accountant_id)
    activities = Activity.objects.filter(tenant=request.user.tenant, is_active=True)
    cities = Client.objects.filter(tenant=request.user.tenant, is_active=True, client_type=Client.TYPE_ACTUAL).exclude(city='').values_list('city', flat=True).distinct().order_by('city')
    sales_users = accountant_users = []
    if request.user.is_admin:
        from accounts.utils import assignable_users
        sales_users = assignable_users(request.user.tenant, UserModel.ROLE_SALES)
        accountant_users = assignable_users(request.user.tenant, UserModel.ROLE_ACCOUNTANT)
    filters = {'q': q, 'city': city, 'district': district, 'activity': activity, 'sales': sales_id, 'accountant': accountant_id}
    from .models import ClientCommissionRule
    commission_rules = {(r.client_id, r.department): r for r in ClientCommissionRule.objects.filter(client__tenant=request.user.tenant)}
    paginator = Paginator(clients.order_by('-created_at'), 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'clients/list.html', {
        'clients': page_obj, 'page_obj': page_obj, 'filters': filters,
        'activities': activities, 'cities': cities,
        'sales_users': sales_users, 'accountant_users': accountant_users,
        'commission_rules': commission_rules,
    })


@login_required
def targeted_list(request):
    from accounts.models import User as UserModel
    clients = get_tenant_clients(request.user, Client.TYPE_POTENTIAL)
    q = request.GET.get('q', '')
    city = request.GET.get('city', '')
    district = request.GET.get('district', '')
    activity = request.GET.get('activity', '')
    sales_id = request.GET.get('sales', '')
    if q:
        clients = clients.filter(name__icontains=q) | clients.filter(company__icontains=q)
    if city:
        clients = clients.filter(city=city)
    if district:
        clients = clients.filter(district__icontains=district)
    if activity:
        clients = clients.filter(activity_id=activity)
    if sales_id and request.user.is_admin:
        clients = clients.filter(assigned_sales_id=sales_id)
    # وسم كل عميل بالأقسام الموجود فيها، وفلترة حسب حالة التحويل
    from django.db.models import Exists, OuterRef, Q
    from workflow.models import ReviewClient
    from zatca.models import ZatcaClient
    clients = clients.annotate(
        in_review=Exists(ReviewClient.objects.filter(source_client=OuterRef('pk'))),
        in_zatca=Exists(ZatcaClient.objects.filter(source_client=OuterRef('pk'))),
    )
    conv = request.GET.get('conv', 'new')
    if conv == 'new':
        clients = clients.filter(in_review=False, in_zatca=False)
    elif conv == 'review':
        clients = clients.filter(in_review=True)
    elif conv == 'zatca':
        clients = clients.filter(in_zatca=True)
    elif conv == 'both':
        clients = clients.filter(in_review=True, in_zatca=True)
    elif conv == 'converted':
        clients = clients.filter(Q(in_review=True) | Q(in_zatca=True))

    activities = Activity.objects.filter(tenant=request.user.tenant, is_active=True)
    cities = Client.objects.filter(tenant=request.user.tenant, is_active=True, client_type=Client.TYPE_POTENTIAL).exclude(city='').values_list('city', flat=True).distinct().order_by('city')
    from accounts.utils import assignable_users
    sales_users = []
    if request.user.is_admin:
        sales_users = assignable_users(request.user.tenant, UserModel.ROLE_SALES)
    total_count = Client.objects.filter(
        tenant=request.user.tenant, client_type=Client.TYPE_POTENTIAL
    ).annotate(
        in_review=Exists(ReviewClient.objects.filter(source_client=OuterRef('pk'))),
        in_zatca=Exists(ZatcaClient.objects.filter(source_client=OuterRef('pk'))),
    ).filter(in_review=False, in_zatca=False).count()
    filters = {'q': q, 'city': city, 'district': district, 'activity': activity,
               'sales': sales_id, 'conv': conv}
    accountant_users = assignable_users(request.user.tenant, UserModel.ROLE_ACCOUNTANT)
    paginator = Paginator(clients.order_by('-created_at'), 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'clients/targeted_list.html', {
        'clients': page_obj, 'page_obj': page_obj, 'filters': filters,
        'activities': activities, 'cities': cities,
        'sales_users': sales_users,
        'accountant_users': accountant_users,
        'total_count': total_count,
    })


@login_required
def client_create(request):
    from .forms import ClientForm
    client_type = request.GET.get('type', Client.TYPE_ACTUAL)
    form = ClientForm(request.POST or None, user=request.user)
    if request.method == 'POST' and form.is_valid():
        client = form.save(commit=False)
        client.tenant = request.user.tenant
        client.created_by = request.user
        client.client_type = client_type
        if request.user.is_sales:
            client.assigned_sales = request.user
        client.save()
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_CREATE, obj=client)
        messages.success(request, 'تم إضافة العميل بنجاح')
        if client_type == Client.TYPE_POTENTIAL:
            return redirect('targeted_list')
        return redirect('clients_list')
    title = 'إضافة عميل مستهدف' if client_type == Client.TYPE_POTENTIAL else 'إضافة عميل'
    return render(request, 'clients/form.html', {'form': form, 'title': title, 'client_type': client_type})


@login_required
def export_targeted_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from django.http import HttpResponse
    from .forms import SAUDI_CITIES

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'العملاء المستهدفون'
    headers = ['الاسم *', 'اسم الشركة', 'المدينة', 'الحي', 'العنوان', 'رقم التواصل', 'البريد الإلكتروني', 'المسئول', 'الوظيفة', 'النشاط', 'ملاحظات']
    ws.append(headers)

    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1a3a5c')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20

    ws_cities = wb.create_sheet(title='المدن')
    city_names = [c[0] for c in SAUDI_CITIES if c[0]]
    for i, city in enumerate(city_names, 1):
        ws_cities.cell(row=i, column=1, value=city)
    ws_cities.sheet_state = 'hidden'

    city_ref = "المدن!" + "$A$1:$A$" + str(len(city_names))
    dv_city = DataValidation(
        type='list',
        formula1=city_ref,
        allow_blank=True,
        showErrorMessage=True,
        error='يرجى اختيار مدينة من القائمة',
        errorTitle='مدينة غير صحيحة',
    )
    ws.add_data_validation(dv_city)
    dv_city.sqref = 'C2:C1000'

    activities = list(Activity.objects.filter(tenant=request.user.tenant, is_active=True).values_list('name', flat=True))
    if activities:
        ws_act = wb.create_sheet(title='الانشطة')
        for i, act in enumerate(activities, 1):
            ws_act.cell(row=i, column=1, value=act)
        ws_act.sheet_state = 'hidden'
        act_ref = "الانشطة!" + "$A$1:$A$" + str(len(activities))
        dv_act = DataValidation(
            type='list',
            formula1=act_ref,
            allow_blank=True,
            showErrorMessage=True,
            error='يرجى اختيار نشاط من القائمة',
            errorTitle='نشاط غير صحيح',
        )
        ws.add_data_validation(dv_act)
        dv_act.sqref = 'J2:J1000'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="targeted_clients_template.xlsx"'
    wb.save(response)
    return response


@login_required
def import_targeted_clients(request):
    import openpyxl
    if request.method == 'POST' and request.FILES.get('excel_file'):
        f = request.FILES['excel_file']
        if f.size > 5 * 1024 * 1024:
            messages.error(request, 'حجم الملف يتجاوز الحد المسموح (5 MB)')
            return redirect('targeted_list')
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            added = 0
            errors = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = str(row[0]).strip() if row[0] else ''
                if not name or name == 'None':
                    continue
                try:
                    Client.objects.create(
                        tenant=request.user.tenant,
                        created_by=request.user,
                        client_type=Client.TYPE_POTENTIAL,
                        assigned_sales=request.user if request.user.is_sales else None,
                        name=name,
                        company=str(row[1] or '').strip(),
                        city=str(row[2] or '').strip(),
                        district=str(row[3] or '').strip(),
                        address=str(row[4] or '').strip(),
                        phone=str(row[5] or '').strip(),
                        email=str(row[6] or '').strip(),
                        responsible_person=str(row[7] or '').strip(),
                        job_title=str(row[8] or '').strip(),
                        notes=str(row[10] or '').strip(),
                    )
                    added += 1
                except Exception as e:
                    errors.append(f'سطر {i}: {e}')
            if added:
                messages.success(request, f'تم استيراد {added} عميل بنجاح')
            if errors:
                for err in errors[:5]:
                    messages.warning(request, err)
        except Exception as e:
            messages.error(request, f'خطأ في قراءة الملف: {e}')
    return redirect('targeted_list')


@login_required
def client_convert(request, pk):
    if request.method != 'POST':
        return redirect('targeted_list')
    from django.utils import timezone
    client = get_object_or_404(Client, pk=pk, tenant=request.user.tenant)
    target = request.POST.get('target', 'actual')

    # منع تكرار التحويل لقسم موجود فيه العميل أصلاً
    if client.is_in_section(target):
        labels = {'actual': 'العملاء الفعليين', 'review': 'قسم المراجعة', 'zatca': 'قسم ZATCA'}
        messages.warning(request, f'"{client.name}" موجود بالفعل في {labels.get(target, target)}')
        return redirect('targeted_list')

    if target == 'zatca':
        from zatca.models import ZatcaClient
        from accounts.models import User as UserModel
        accountant_id = request.POST.get('accountant_id')
        accountant = None
        if accountant_id:
            from accounts.utils import assignable_users
            accountant = assignable_users(request.user.tenant, UserModel.ROLE_ACCOUNTANT).filter(pk=accountant_id).first()
        period = request.POST.get('period_months', '1')
        period = int(period) if period in ('1', '3', '6', '12') else 1
        zatca_client = ZatcaClient.objects.create(
            tenant=client.tenant,
            source_client=client,
            name=client.name,
            company=client.company,
            phone=client.phone,
            email=client.email,
            city=client.city,
            district=client.district,
            address=client.address,
            responsible_person=client.responsible_person,
            job_title=client.job_title,
            notes=client.notes,
            distinguished_number=client.distinguished_number,
            secret_number=client.secret_number,
            assigned_accountant=accountant,
            period_months=period,
            created_by=request.user,
        )
        client.converted_status = 'zatca'
        client.converted_at = timezone.now()
        client.save(update_fields=['converted_status', 'converted_at'])
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_CREATE, obj=zatca_client)
        log_action(request, AuditLog.ACTION_UPDATE, obj=client,
                   changes={'تحويل': {'من': 'مستهدف', 'إلى': 'ZATCA'}})
        messages.success(request, f'تم تحويل "{client.name}" إلى قسم ZATCA')
        return redirect('zatca_detail', pk=zatca_client.pk)

    elif target == 'review':
        from workflow.models import ReviewClient
        from workflow.views import _create_stages
        review_client = ReviewClient.objects.create(
            tenant=client.tenant,
            source_client=client,
            name=client.name,
            company=client.company,
            phone=client.phone,
            email=client.email,
            city=client.city,
            district=client.district,
            address=client.address,
            responsible_person=client.responsible_person,
            job_title=client.job_title,
            notes=client.notes,
            distinguished_number=client.distinguished_number,
            secret_number=client.secret_number,
            assigned_reviewer=client.assigned_sales,
            created_by=request.user,
        )
        _create_stages(review_client, timezone.localdate(), {})
        client.converted_status = Client.CONVERTED_REVIEW
        client.converted_at = timezone.now()
        client.save(update_fields=['converted_status', 'converted_at'])
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_UPDATE, obj=client,
                   changes={'تحويل': {'من': 'مستهدف', 'إلى': 'قسم المراجعة'}})
        messages.success(request, f'تم تحويل "{client.name}" إلى قسم المراجعة')
        return redirect('workflow_list')
    else:
        client.client_type = Client.TYPE_ACTUAL
        client.converted_status = Client.CONVERTED_ACTUAL
        client.converted_at = timezone.now()
        client.save(update_fields=['client_type', 'converted_status', 'converted_at'])
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_UPDATE, obj=client,
                   changes={'تحويل': {'من': 'مستهدف', 'إلى': 'عميل فعلي'}})
        messages.success(request, f'تم تحويل "{client.name}" إلى عميل فعلي')
        return redirect('targeted_list')


@login_required
@admin_required
def activities_list(request):
    activities = Activity.objects.filter(tenant=request.user.tenant)
    return render(request, 'clients/activities.html', {'activities': activities})


@login_required
@admin_required
def activity_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            Activity.objects.create(tenant=request.user.tenant, name=name, created_by=request.user)
            messages.success(request, f'تم إضافة النشاط "{name}"')
        else:
            messages.error(request, 'اسم النشاط مطلوب')
    return redirect('activities_list')


@login_required
@admin_required
def activity_toggle(request, pk):
    if request.method != 'POST':
        return redirect('activities_list')
    activity = get_object_or_404(Activity, pk=pk, tenant=request.user.tenant)
    activity.is_active = not activity.is_active
    activity.save()
    return redirect('activities_list')


@login_required
@admin_required
def toggle_commissionable(request, pk):
    from django.http import JsonResponse, HttpResponseNotAllowed
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    client = get_object_or_404(Client, pk=pk, tenant=request.user.tenant)
    if not can_access_client(request.user, client):
        return JsonResponse({'status': 'forbidden'}, status=403)
    client.is_commissionable = not client.is_commissionable
    client.save(update_fields=['is_commissionable'])
    return JsonResponse({'status': 'ok', 'is_commissionable': client.is_commissionable})


@login_required
def client_detail(request, pk):
    from .models import ClientNote, ClientAttachment
    client = get_object_or_404(Client, pk=pk, tenant=request.user.tenant)
    if not can_access_client(request.user, client):
        messages.error(request, 'ليس لديك صلاحية الوصول لهذا العميل')
        return redirect('clients_list')
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'upload' and request.FILES.get('file'):
            f = request.FILES['file']
            from core.uploads import validate_upload
            ok, err = validate_upload(f)
            if not ok:
                messages.error(request, err)
                return redirect('client_detail', pk=pk)
            att = ClientAttachment.objects.create(client=client, file=f, name=f.name, uploaded_by=request.user)
            from audit_log.utils import log_action
            from audit_log.models import AuditLog
            log_action(request, AuditLog.ACTION_CREATE, obj=att,
                       model_name='ClientAttachment', object_repr=f.name)
            messages.success(request, 'تم رفع الملف')
        elif request.POST.get('note'):
            note = ClientNote.objects.create(client=client, note=request.POST['note'], created_by=request.user)
            from audit_log.utils import log_action
            from audit_log.models import AuditLog
            log_action(request, AuditLog.ACTION_CREATE, obj=note,
                       model_name='ClientNote', object_repr=str(client))
            messages.success(request, 'تم إضافة الملاحظة')
        return redirect('client_detail', pk=pk)
    from accounts.models import User as UserModel
    from accounts.utils import assignable_users
    accountant_users = assignable_users(request.user.tenant, UserModel.ROLE_ACCOUNTANT)
    return render(request, 'clients/detail.html', {'client': client, 'accountant_users': accountant_users})


@login_required
def attachment_delete(request, pk):
    from .models import ClientAttachment
    import os
    att = get_object_or_404(ClientAttachment, pk=pk, client__tenant=request.user.tenant)
    client_pk = att.client.pk
    if request.method == 'POST':
        if att.file and os.path.isfile(att.file.path):
            os.remove(att.file.path)
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_DELETE, model_name='ClientAttachment',
                   object_repr=att.name, object_id=str(att.pk))
        att.delete()
        messages.success(request, 'تم حذف المرفق')
    return redirect('client_detail', pk=client_pk)


@login_required
def client_edit(request, pk):
    from .forms import ClientForm
    client = get_object_or_404(Client, pk=pk, tenant=request.user.tenant)
    if not can_access_client(request.user, client):
        messages.error(request, 'ليس لديك صلاحية تعديل هذا العميل')
        return redirect('clients_list')
    form = ClientForm(request.POST or None, instance=client, user=request.user)
    if request.method == 'POST' and form.is_valid():
        form.save()
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_UPDATE, obj=client)
        messages.success(request, 'تم تحديث بيانات العميل')
        return redirect('client_detail', pk=pk)
    return render(request, 'clients/form.html', {'form': form, 'title': 'تعديل عميل', 'client_type': client.client_type})


@login_required
@admin_required
def commission_rule_save(request, client_pk):
    from .models import ClientCommissionRule
    client = get_object_or_404(Client, pk=client_pk, tenant=request.user.tenant)
    department = request.POST.get('department')
    if request.method == 'POST' and department:
        ClientCommissionRule.objects.update_or_create(
            client=client,
            department=department,
            defaults={
                'commission_type': request.POST.get('commission_type', 'fixed'),
                'value': request.POST.get('value') or 0,
                'max_amount': request.POST.get('max_amount') or None,
                'reference_note': request.POST.get('reference_note', ''),
                'notes': request.POST.get('notes', ''),
                'created_by': request.user,
            }
        )
        from commissions.models import CommissionEntry
        qs = CommissionEntry.objects.filter(client=client)
        sheet_pk = request.POST.get('sheet_pk')
        if sheet_pk:
            qs = qs.filter(sheet_id=sheet_pk)
        for entry in qs:
            entry.save()
        messages.success(request, 'تم حفظ قاعدة العمولة')
    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER', '')
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = 'clients_list'
    return redirect(next_url)


@login_required
@admin_required
def commission_rule_delete(request, client_pk, department):
    if request.method != 'POST':
        return redirect('clients_list')
    from .models import ClientCommissionRule
    client = get_object_or_404(Client, pk=client_pk, tenant=request.user.tenant)
    ClientCommissionRule.objects.filter(client=client, department=department).delete()
    from commissions.models import CommissionEntry
    qs = CommissionEntry.objects.filter(client=client)
    sheet_pk = request.POST.get('sheet_pk')
    if sheet_pk:
        qs = qs.filter(sheet_id=sheet_pk)
    for entry in qs:
        entry.save()
    messages.success(request, 'تم حذف قاعدة العمولة')
    return redirect('client_detail', pk=client_pk)


@login_required
@admin_required
def client_delete(request, pk):
    if request.method != 'POST':
        return redirect('clients_list')
    client = get_object_or_404(Client, pk=pk, tenant=request.user.tenant)
    name = client.name
    client_type = client.client_type
    from audit_log.utils import log_action
    from audit_log.models import AuditLog
    log_action(request, AuditLog.ACTION_DELETE, model_name='Client',
               object_repr=name, object_id=str(pk))
    client.delete()
    messages.success(request, f'تم حذف العميل "{name}" بنجاح')
    if client_type == Client.TYPE_POTENTIAL:
        return redirect('targeted_list')
    return redirect('clients_list')
