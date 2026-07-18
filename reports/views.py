from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.utils import timezone
from accounts.decorators import admin_required


def _date_filters(request):
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    return date_from, date_to


def _apply_commission_filters(entries, date_from, date_to, role_filter, user_filter):
    """Apply all commissions filters to an entries queryset. Used by view + exports."""
    if date_from:
        entries = entries.filter(sheet__created_at__date__gte=date_from)
    if date_to:
        entries = entries.filter(sheet__created_at__date__lte=date_to)

    if role_filter == 'sales':
        if user_filter:
            entries = entries.filter(sales_rep_id=user_filter)
        else:
            entries = entries.filter(commission_amount__gt=0)
    elif role_filter == 'accountant':
        if user_filter:
            entries = entries.filter(client__assigned_accountant_id=user_filter)
        else:
            entries = entries.filter(accountant_commission_amount__gt=0)
    elif role_filter == 'review':
        if user_filter:
            entries = entries.filter(client__assigned_review_id=user_filter)
        else:
            entries = entries.filter(review_commission_amount__gt=0)

    return entries


def _build_bysheet(sheets_qs, entry_ids):
    """Annotate sheets with totals for the given filtered entry_ids."""
    return (
        sheets_qs.filter(entries__id__in=entry_ids)
        .annotate(
            t_amount     = Sum('entries__amount',                       filter=Q(entries__id__in=entry_ids)),
            t_sales      = Sum('entries__commission_amount',            filter=Q(entries__id__in=entry_ids)),
            t_accountant = Sum('entries__accountant_commission_amount', filter=Q(entries__id__in=entry_ids)),
            t_review     = Sum('entries__review_commission_amount',     filter=Q(entries__id__in=entry_ids)),
        )
        .distinct()
        .order_by('-created_at')
    )


# ─────────────────────────────────────────
# الصفحة الرئيسية للتقارير
# ─────────────────────────────────────────
@login_required
@admin_required
def reports_home(request):
    from clients.models import Client
    from commissions.models import CommissionEntry
    from calendar_app.models import Event
    from workflow.models import ReviewClient, WorkflowStage

    tenant = request.user.tenant
    today  = timezone.localdate()

    ctx = {
        'total_clients':      Client.objects.filter(tenant=tenant, is_active=True, client_type='actual').count(),
        'total_targeted':     Client.objects.filter(tenant=tenant, is_active=True, client_type='potential').count(),
        'total_commission':   CommissionEntry.objects.filter(sheet__tenant=tenant).aggregate(
                                  s=Sum('commission_amount'))['s'] or 0,
        'total_accountant':   CommissionEntry.objects.filter(sheet__tenant=tenant).aggregate(
                                  s=Sum('accountant_commission_amount'))['s'] or 0,
        'total_review_comm':  CommissionEntry.objects.filter(sheet__tenant=tenant).aggregate(
                                  s=Sum('review_commission_amount'))['s'] or 0,
        'total_amount':       CommissionEntry.objects.filter(sheet__tenant=tenant).aggregate(
                                  s=Sum('amount'))['s'] or 0,
        'events_done':        Event.objects.filter(tenant=tenant, status='done').count(),
        'events_cancelled':   Event.objects.filter(tenant=tenant, status='cancelled').count(),
        'events_pending':     Event.objects.filter(tenant=tenant, status='pending').count(),
        'workflow_clients':   ReviewClient.objects.filter(tenant=tenant).count(),
        'workflow_overdue':   WorkflowStage.objects.filter(
                                  client__tenant=tenant,
                                  due_date__lt=today,
                                  status__in=['in_progress', 'waiting_approval']
                              ).count(),
        'today': today,
    }
    return render(request, 'reports/home.html', ctx)


# ─────────────────────────────────────────
# تقرير العملاء
# ─────────────────────────────────────────
@login_required
@admin_required
def report_clients(request):
    from clients.models import Client, Activity
    from accounts.models import User
    from workflow.models import ReviewClient

    tenant = request.user.tenant
    date_from, date_to = _date_filters(request)
    city_filter  = request.GET.get('city', '')
    sales_filter = request.GET.get('sales', '')

    # العملاء الفعليون
    qs = Client.objects.filter(tenant=tenant, is_active=True, client_type='actual')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if city_filter:
        qs = qs.filter(city=city_filter)
    if sales_filter:
        qs = qs.filter(assigned_sales_id=sales_filter)

    # عملاء المراجعة
    review_qs = ReviewClient.objects.filter(tenant=tenant)
    if date_from:
        review_qs = review_qs.filter(created_at__date__gte=date_from)
    if date_to:
        review_qs = review_qs.filter(created_at__date__lte=date_to)
    if city_filter:
        review_qs = review_qs.filter(city=city_filter)

    # عملاء ZATCA
    from zatca.models import ZatcaClient
    zatca_qs = ZatcaClient.objects.filter(tenant=tenant)
    if date_from:
        zatca_qs = zatca_qs.filter(created_at__date__gte=date_from)
    if date_to:
        zatca_qs = zatca_qs.filter(created_at__date__lte=date_to)
    if city_filter:
        zatca_qs = zatca_qs.filter(city=city_filter)

    by_city     = qs.values('city').annotate(count=Count('id')).order_by('-count')
    by_activity = qs.values('activity__name').annotate(count=Count('id')).order_by('-count')
    by_sales    = qs.values('assigned_sales__first_name','assigned_sales__last_name','assigned_sales__username').annotate(count=Count('id')).order_by('-count')
    cities      = sorted(set(
        Client.objects.filter(tenant=tenant, is_active=True, client_type='actual').exclude(city='').values_list('city', flat=True)
    ))
    sales_users = User.objects.filter(tenant=tenant, role='sales', is_active=True)

    return render(request, 'reports/clients.html', {
        'clients': qs.select_related('assigned_sales','assigned_accountant','activity'),
        'review_clients': review_qs.select_related('assigned_reviewer'),
        'zatca_clients': zatca_qs.select_related('assigned_accountant'),
        'total': qs.count() + review_qs.count() + zatca_qs.count(),
        'total_actual': qs.count(),
        'total_review': review_qs.count(),
        'total_zatca': zatca_qs.count(),
        'by_city': by_city,
        'by_activity': by_activity,
        'by_sales': by_sales,
        'cities': cities,
        'sales_users': sales_users,
        'filters': {'date_from': date_from, 'date_to': date_to,
                    'city': city_filter, 'sales': sales_filter},
    })


# ─────────────────────────────────────────
# تقرير العمولات
# ─────────────────────────────────────────
@login_required
@admin_required
def report_commissions(request):
    from commissions.models import CommissionSheet, CommissionEntry
    from accounts.models import User

    tenant      = request.user.tenant
    date_from, date_to = _date_filters(request)
    role_filter = request.GET.get('role', '')
    user_filter = request.GET.get('user', '')

    sheets  = CommissionSheet.objects.filter(tenant=tenant)
    entries = CommissionEntry.objects.filter(sheet__tenant=tenant)
    entries = _apply_commission_filters(entries, date_from, date_to, role_filter, user_filter)

    from accounts.utils import assignable_users
    sales_users      = assignable_users(tenant, User.ROLE_SALES)
    accountant_users = assignable_users(tenant, User.ROLE_ACCOUNTANT)
    review_users     = assignable_users(tenant, User.ROLE_REVIEW)

    totals = entries.aggregate(
        total_amount    = Sum('amount'),
        total_sales     = Sum('commission_amount'),
        total_accountant= Sum('accountant_commission_amount'),
        total_review    = Sum('review_commission_amount'),
    )

    entry_ids = list(entries.values_list('id', flat=True))
    by_sheet  = _build_bysheet(sheets, entry_ids)

    top_clients = entries.values('client__name').annotate(
        total=Sum('amount')
    ).order_by('-total')[:10]

    return render(request, 'reports/commissions.html', {
        'totals':           totals,
        'by_sheet':         by_sheet,
        'sales_users':      sales_users,
        'accountant_users': accountant_users,
        'review_users':     review_users,
        'top_clients':      top_clients,
        'filters': {
            'date_from': date_from, 'date_to': date_to,
            'role': role_filter, 'user': user_filter,
        },
    })


# ─────────────────────────────────────────
# تقرير الأحداث
# ─────────────────────────────────────────
@login_required
@admin_required
def report_events(request):
    from calendar_app.models import Event
    from accounts.models import User

    tenant = request.user.tenant
    date_from, date_to = _date_filters(request)
    user_filter   = request.GET.get('user', '')
    source_filter = request.GET.get('source', '')

    qs = Event.objects.filter(tenant=tenant)
    if date_from:
        qs = qs.filter(start_datetime__date__gte=date_from)
    if date_to:
        qs = qs.filter(start_datetime__date__lte=date_to)
    if user_filter:
        qs = qs.filter(assigned_to_id=user_filter)
    if source_filter:
        qs = qs.filter(source=source_filter)

    by_status = qs.values('status').annotate(count=Count('id'))
    by_type   = qs.values('event_type').annotate(count=Count('id'))
    by_user   = qs.values('assigned_to__first_name','assigned_to__last_name','assigned_to__username').annotate(
                    total=Count('id'), done=Count('id', filter=Q(status='done'))
                ).order_by('-total')
    users = User.objects.filter(tenant=tenant, is_active=True)

    return render(request, 'reports/events.html', {
        'total': qs.count(),
        'by_status': by_status,
        'by_type': by_type,
        'by_user': by_user,
        'events': qs.select_related('assigned_to','client').order_by('-start_datetime')[:200],
        'users': users,
        'filters': {'date_from': date_from, 'date_to': date_to,
                    'user': user_filter, 'source': source_filter},
    })


# ─────────────────────────────────────────
# تقرير قسم المراجعة
# ─────────────────────────────────────────
@login_required
@admin_required
def report_workflow(request):
    from workflow.models import ReviewClient, WorkflowStage

    tenant       = request.user.tenant
    today        = timezone.localdate()
    stage_filter = request.GET.get('stage', '')
    status_filter= request.GET.get('status', '')

    clients = ReviewClient.objects.filter(tenant=tenant).prefetch_related('stages')

    if stage_filter or status_filter:
        stage_qs = WorkflowStage.objects.filter(client__tenant=tenant)
        if stage_filter:
            stage_qs = stage_qs.filter(stage=stage_filter)
        if status_filter:
            stage_qs = stage_qs.filter(status=status_filter)
        client_ids = stage_qs.values_list('client_id', flat=True).distinct()
        clients = clients.filter(pk__in=client_ids)

    by_stage = WorkflowStage.objects.filter(client__tenant=tenant).values('stage','status').annotate(count=Count('id'))
    overdue  = WorkflowStage.objects.filter(
        client__tenant=tenant, due_date__lt=today,
        status__in=['in_progress','waiting_approval']
    ).select_related('client')

    return render(request, 'reports/workflow.html', {
        'clients':       clients,
        'by_stage':      by_stage,
        'overdue':       overdue,
        'stage_choices': WorkflowStage.STAGE_CHOICES,
        'status_choices':WorkflowStage.STATUS_CHOICES,
        'today':         today,
        'filters':       {'stage': stage_filter, 'status': status_filter},
    })


# ─────────────────────────────────────────
# تقرير المستخدمين
# ─────────────────────────────────────────
@login_required
@admin_required
def report_users(request):
    from accounts.models import User
    from calendar_app.models import Event

    tenant = request.user.tenant
    date_from, date_to = _date_filters(request)

    users = User.objects.filter(tenant=tenant, is_active=True)

    events_qs = Event.objects.filter(tenant=tenant)
    if date_from:
        events_qs = events_qs.filter(start_datetime__date__gte=date_from)
    if date_to:
        events_qs = events_qs.filter(start_datetime__date__lte=date_to)

    stats = {
        row['assigned_to']: row
        for row in events_qs.values('assigned_to').annotate(
            total=Count('id'),
            done=Count('id', filter=Q(status='done')),
            cancelled=Count('id', filter=Q(status='cancelled')),
            pending=Count('id', filter=Q(status='pending')),
        )
    }

    data = [
        {
            'user':      u,
            'total':     stats.get(u.pk, {}).get('total', 0),
            'done':      stats.get(u.pk, {}).get('done', 0),
            'cancelled': stats.get(u.pk, {}).get('cancelled', 0),
            'pending':   stats.get(u.pk, {}).get('pending', 0),
        }
        for u in users
    ]

    return render(request, 'reports/users.html', {
        'data': data,
        'filters': {'date_from': date_from, 'date_to': date_to},
    })


# ═══════════════════════════════════════════
# تصدير Excel
# ═══════════════════════════════════════════
def _make_excel_response(filename):
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(ws, headers, fill_color='1a3a5c'):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor=fill_color)
        cell.alignment = Alignment(horizontal='center')


@login_required
@admin_required
def export_clients_excel(request):
    import openpyxl
    from clients.models import Client
    from workflow.models import ReviewClient

    tenant       = request.user.tenant
    date_from    = request.GET.get('date_from', '')
    date_to      = request.GET.get('date_to', '')
    city_filter  = request.GET.get('city', '')
    sales_filter = request.GET.get('sales', '')

    qs = Client.objects.filter(tenant=tenant, is_active=True, client_type='actual').select_related('assigned_sales','assigned_accountant','activity')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if city_filter:
        qs = qs.filter(city=city_filter)
    if sales_filter:
        qs = qs.filter(assigned_sales_id=sales_filter)

    review_qs = ReviewClient.objects.filter(tenant=tenant).select_related('assigned_reviewer')
    if date_from:
        review_qs = review_qs.filter(created_at__date__gte=date_from)
    if date_to:
        review_qs = review_qs.filter(created_at__date__lte=date_to)
    if city_filter:
        review_qs = review_qs.filter(city=city_filter)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'العملاء الفعليون'
    _style_header(ws, ['الاسم','الشركة','المدينة','الجوال','المندوب','المحاسب','النشاط','تاريخ الإضافة'])
    for c in qs:
        ws.append([
            c.name, c.company, c.city, c.phone,
            c.assigned_sales.get_full_name() if c.assigned_sales else '',
            c.assigned_accountant.get_full_name() if c.assigned_accountant else '',
            c.activity.name if c.activity else '',
            str(c.created_at.date()) if c.created_at else '',
        ])

    ws2 = wb.create_sheet(title='قسم المراجعة')
    _style_header(ws2, ['الاسم','الشركة','المدينة','الجوال','المراجع','تاريخ الإضافة'])
    for c in review_qs:
        ws2.append([
            c.name, c.company, c.city, c.phone,
            c.assigned_reviewer.get_full_name() if c.assigned_reviewer else '',
            str(c.created_at.date()) if c.created_at else '',
        ])

    from zatca.models import ZatcaClient
    zatca_qs2 = ZatcaClient.objects.filter(tenant=tenant).select_related('assigned_accountant')
    if date_from:
        zatca_qs2 = zatca_qs2.filter(created_at__date__gte=date_from)
    if date_to:
        zatca_qs2 = zatca_qs2.filter(created_at__date__lte=date_to)
    if city_filter:
        zatca_qs2 = zatca_qs2.filter(city=city_filter)

    ws3 = wb.create_sheet(title='ZATCA')
    _style_header(ws3, ['الاسم','الشركة','المدينة','الجوال','المحاسب','الحالة','تاريخ الإضافة'], fill_color='b45309')
    for c in zatca_qs2:
        ws3.append([
            c.name, c.company, c.city, c.phone,
            c.assigned_accountant.get_full_name() if c.assigned_accountant else '',
            c.get_status_display(),
            str(c.created_at.date()) if c.created_at else '',
        ])

    response = _make_excel_response('clients.xlsx')
    wb.save(response)
    return response


@login_required
@admin_required
def export_commissions_excel(request):
    import openpyxl
    from commissions.models import CommissionEntry

    tenant      = request.user.tenant
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    role_filter = request.GET.get('role', '')
    user_filter = request.GET.get('user', '')

    entries = CommissionEntry.objects.filter(sheet__tenant=tenant).select_related('client','sheet')
    entries = _apply_commission_filters(entries, date_from, date_to, role_filter, user_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'العمولات'
    _style_header(ws, ['الشيت','العميل','المبلغ','عمولة المناديب','عمولة المحاسبين','عمولة المراجعين'])
    for e in entries:
        ws.append([
            str(e.sheet), e.client.name if e.client else '',
            e.amount, e.commission_amount,
            e.accountant_commission_amount, e.review_commission_amount,
        ])
    response = _make_excel_response('commissions.xlsx')
    wb.save(response)
    return response


@login_required
@admin_required
def export_events_excel(request):
    import openpyxl
    from calendar_app.models import Event

    tenant        = request.user.tenant
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')
    user_filter   = request.GET.get('user', '')
    source_filter = request.GET.get('source', '')

    qs = Event.objects.filter(tenant=tenant).select_related('assigned_to','client')
    if date_from:
        qs = qs.filter(start_datetime__date__gte=date_from)
    if date_to:
        qs = qs.filter(start_datetime__date__lte=date_to)
    if user_filter:
        qs = qs.filter(assigned_to_id=user_filter)
    if source_filter:
        qs = qs.filter(source=source_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الأحداث'
    _style_header(ws, ['العنوان','النوع','الحالة','العميل','المسند إلى','التاريخ','ملاحظات'])
    for e in qs:
        ws.append([
            e.title, e.get_event_type_display(), e.get_status_display(),
            e.client.name if e.client else '',
            e.assigned_to.get_full_name() if e.assigned_to else '',
            str(e.start_datetime.date()),
            e.notes,
        ])
    response = _make_excel_response('events.xlsx')
    wb.save(response)
    return response


@login_required
@admin_required
def export_workflow_excel(request):
    import openpyxl
    from workflow.models import ReviewClient, WorkflowStage

    tenant       = request.user.tenant
    stage_filter = request.GET.get('stage', '')
    status_filter= request.GET.get('status', '')

    clients = ReviewClient.objects.filter(tenant=tenant).prefetch_related('stages')
    if stage_filter or status_filter:
        stage_qs = WorkflowStage.objects.filter(client__tenant=tenant)
        if stage_filter:
            stage_qs = stage_qs.filter(stage=stage_filter)
        if status_filter:
            stage_qs = stage_qs.filter(status=status_filter)
        client_ids = stage_qs.values_list('client_id', flat=True).distinct()
        clients = clients.filter(pk__in=client_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'قسم المراجعة'
    _style_header(ws, ['العميل','الشركة','المرحلة','الحالة','تاريخ البداية','تاريخ النهاية','تاريخ الاستحقاق','المدة (يوم)'])
    from django.utils import timezone
    today = timezone.localdate()
    for client in clients:
        for stage in client.stages.all():
            days = stage.days_in_stage
            ws.append([
                client.name, client.company,
                stage.get_stage_display(), stage.get_status_display(),
                str(stage.start_date) if stage.start_date else '',
                str(stage.end_date)   if stage.end_date   else '',
                str(stage.due_date)   if stage.due_date   else '',
                days if days is not None else '',
            ])
    response = _make_excel_response('workflow.xlsx')
    wb.save(response)
    return response


def _render_pdf(title, headers, rows, filename, username, has_totals=False):
    """Render a PDF using WeasyPrint from pdf_table.html template."""
    import weasyprint
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    from django.utils import timezone

    def _is_num(val):
        try:
            float(str(val).replace(',', ''))
            return str(val) not in ('', '—')
        except (ValueError, TypeError):
            return False

    now_str = timezone.localtime().strftime('%Y/%m/%d %H:%M')
    table_rows = [
        [{'value': cell, 'is_num': _is_num(cell)} for cell in row]
        for row in rows
    ]
    html = render_to_string('pdf_table.html', {
        'title': title,
        'headers': headers,
        'rows': table_rows,
        'username': username,
        'now': now_str,
        'has_totals': has_totals,
    })
    pdf = weasyprint.HTML(string=html, base_url='/').write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
@admin_required
def export_clients_pdf(request):
    from clients.models import Client
    from workflow.models import ReviewClient

    tenant       = request.user.tenant
    date_from    = request.GET.get('date_from', '')
    date_to      = request.GET.get('date_to', '')
    city_filter  = request.GET.get('city', '')
    sales_filter = request.GET.get('sales', '')

    clients = Client.objects.filter(tenant=tenant, is_active=True, client_type='actual').select_related('assigned_sales', 'assigned_accountant', 'activity')
    if date_from:
        clients = clients.filter(created_at__date__gte=date_from)
    if date_to:
        clients = clients.filter(created_at__date__lte=date_to)
    if city_filter:
        clients = clients.filter(city=city_filter)
    if sales_filter:
        clients = clients.filter(assigned_sales_id=sales_filter)

    review_qs = ReviewClient.objects.filter(tenant=tenant).select_related('assigned_reviewer')
    if date_from:
        review_qs = review_qs.filter(created_at__date__gte=date_from)
    if date_to:
        review_qs = review_qs.filter(created_at__date__lte=date_to)
    if city_filter:
        review_qs = review_qs.filter(city=city_filter)

    username = request.user.get_full_name() or request.user.username
    headers = ['الاسم', 'الشركة', 'المدينة', 'الجوال', 'النوع', 'المندوب', 'المحاسب', 'المراجع', 'تاريخ الإضافة']
    rows = []
    for c in clients:
        rows.append([
            c.name or '', c.company or '', c.city or '', c.phone or '',
            'فعلي',
            c.assigned_sales.get_full_name() if c.assigned_sales else '—',
            c.assigned_accountant.get_full_name() if c.assigned_accountant else '—',
            '—',
            str(c.created_at.date()) if c.created_at else '',
        ])
    for c in review_qs:
        rows.append([
            c.name or '', c.company or '', c.city or '', c.phone or '',
            'مراجعة',
            '—', '—',
            c.assigned_reviewer.get_full_name() if c.assigned_reviewer else '—',
            str(c.created_at.date()) if c.created_at else '',
        ])

    from zatca.models import ZatcaClient
    zatca_qs = ZatcaClient.objects.filter(tenant=tenant).select_related('assigned_accountant')
    if date_from:
        zatca_qs = zatca_qs.filter(created_at__date__gte=date_from)
    if date_to:
        zatca_qs = zatca_qs.filter(created_at__date__lte=date_to)
    if city_filter:
        zatca_qs = zatca_qs.filter(city=city_filter)

    for c in zatca_qs:
        rows.append([
            c.name or '', c.company or '', c.city or '', c.phone or '',
            'ZATCA',
            '—',
            c.assigned_accountant.get_full_name() if c.assigned_accountant else '—',
            '—',
            str(c.created_at.date()) if c.created_at else '',
        ])
    return _render_pdf('تقرير العملاء', headers, rows, 'clients.pdf', username)


@login_required
@admin_required
def export_commissions_pdf(request):
    from commissions.models import CommissionEntry
    from collections import OrderedDict

    tenant      = request.user.tenant
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    role_filter = request.GET.get('role', '')
    user_filter = request.GET.get('user', '')

    entries = CommissionEntry.objects.filter(sheet__tenant=tenant).select_related('sheet')
    entries = _apply_commission_filters(entries, date_from, date_to, role_filter, user_filter)
    username = request.user.get_full_name() or request.user.username

    def _fmt(v):
        return f"{v:,.2f}"

    from decimal import Decimal
    zero = Decimal(0)
    sheet_totals = OrderedDict()
    for e in entries:
        sid = e.sheet_id
        if sid not in sheet_totals:
            sheet_totals[sid] = {'name': str(e.sheet), 'amount': zero, 'sales': zero, 'accountant': zero, 'review': zero}
        sheet_totals[sid]['amount']     += e.amount or zero
        sheet_totals[sid]['sales']      += e.commission_amount or zero
        sheet_totals[sid]['accountant'] += e.accountant_commission_amount or zero
        sheet_totals[sid]['review']     += e.review_commission_amount or zero

    headers = ['الشيت', 'إجمالي المبالغ', 'عمولات المناديب', 'عمولات المحاسبين', 'عمولات المراجعين']
    rows = []
    for data in sheet_totals.values():
        rows.append([data['name'], _fmt(data['amount']), _fmt(data['sales']), _fmt(data['accountant']), _fmt(data['review'])])
    return _render_pdf('تقرير العمولات', headers, rows, 'commissions.pdf', username)


@login_required
@admin_required
def export_events_pdf(request):
    from calendar_app.models import Event

    tenant        = request.user.tenant
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')
    user_filter   = request.GET.get('user', '')
    source_filter = request.GET.get('source', '')

    events = Event.objects.filter(tenant=tenant).select_related('assigned_to', 'client').order_by('-start_datetime')
    if date_from:
        events = events.filter(start_datetime__date__gte=date_from)
    if date_to:
        events = events.filter(start_datetime__date__lte=date_to)
    if user_filter:
        events = events.filter(assigned_to_id=user_filter)
    if source_filter:
        events = events.filter(source=source_filter)
    username = request.user.get_full_name() or request.user.username

    headers = ['العنوان', 'النوع', 'الحالة', 'العميل', 'المسند إلى', 'التاريخ', 'ملاحظات']
    rows = []
    for e in events:
        rows.append([
            e.title, e.get_event_type_display(), e.get_status_display(),
            e.client.name if e.client else '—',
            e.assigned_to.get_full_name() if e.assigned_to else '—',
            str(e.start_datetime.date()),
            (e.notes or '')[:60],
        ])
    return _render_pdf('تقرير الأحداث', headers, rows, 'events.pdf', username)


@login_required
@admin_required
def export_workflow_pdf(request):
    from workflow.models import ReviewClient, WorkflowStage

    tenant        = request.user.tenant
    stage_filter  = request.GET.get('stage', '')
    status_filter = request.GET.get('status', '')

    clients = ReviewClient.objects.filter(tenant=tenant).prefetch_related('stages')
    if stage_filter or status_filter:
        stage_qs = WorkflowStage.objects.filter(client__tenant=tenant)
        if stage_filter:
            stage_qs = stage_qs.filter(stage=stage_filter)
        if status_filter:
            stage_qs = stage_qs.filter(status=status_filter)
        client_ids = stage_qs.values_list('client_id', flat=True).distinct()
        clients = clients.filter(pk__in=client_ids)

    username = request.user.get_full_name() or request.user.username
    headers = ['العميل', 'الشركة', 'المرحلة', 'الحالة', 'بداية', 'نهاية', 'استحقاق', 'المدة (يوم)']
    rows = []
    for client in clients:
        for stage in client.stages.all():
            rows.append([
                client.name, client.company or '',
                stage.get_stage_display(), stage.get_status_display(),
                str(stage.start_date) if stage.start_date else '—',
                str(stage.end_date)   if stage.end_date   else '—',
                str(stage.due_date)   if stage.due_date   else '—',
                str(stage.days_in_stage) if stage.days_in_stage is not None else '—',
            ])
    return _render_pdf('تقرير المراجعة', headers, rows, 'workflow.pdf', username)
