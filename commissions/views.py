from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from decimal import Decimal
from .models import CommissionSheet, CommissionEntry, EntryCommissionRule
from clients.models import Client
from accounts.decorators import accountant_required, admin_required
from .forms import CommissionSheetForm


@login_required
@admin_required
def commissions_list(request):
    sheets = CommissionSheet.objects.filter(tenant=request.user.tenant)
    for sheet in sheets:
        agg = sheet.entries.aggregate(
            t_amount=Sum('amount'),
            t_sales=Sum('commission_amount'),
            t_accountant=Sum('accountant_commission_amount'),
            t_review=Sum('review_commission_amount'),
        )
        sheet.total_amount = agg['t_amount'] or 0
        sheet.total_commission = (agg['t_sales'] or 0) + (agg['t_accountant'] or 0) + (agg['t_review'] or 0)
    return render(request, 'commissions/list.html', {'sheets': sheets})


@login_required
@admin_required
def commission_create(request):
    form = CommissionSheetForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        from django.db import transaction
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        with transaction.atomic():
            sheet = form.save(commit=False)
            sheet.tenant = request.user.tenant
            sheet.created_by = request.user
            sheet.save()

            actual_clients = Client.objects.filter(
                tenant=request.user.tenant,
                is_commissionable=True,
                is_active=True,
            ).select_related('assigned_sales')

            entries = []
            for client in actual_clients:
                entries.append(CommissionEntry(
                    sheet=sheet,
                    client=client,
                    sales_rep=client.assigned_sales or request.user,
                    amount=0,
                ))
            if entries:
                CommissionEntry.objects.bulk_create(entries)

        log_action(request, AuditLog.ACTION_CREATE, obj=sheet,
                   changes={'entries_count': {'إلى': str(len(entries))}})
        messages.success(request, f'تم إنشاء الشيت وتعبئته بـ {len(entries)} عميل فعلي')
        return redirect('commission_detail', pk=sheet.pk)
    return render(request, 'commissions/form.html', {'form': form})


@login_required
@admin_required
def commission_detail(request, pk):
    sheet = get_object_or_404(CommissionSheet, pk=pk, tenant=request.user.tenant)
    entries = sheet.entries.select_related('client', 'client__assigned_accountant', 'sales_rep').prefetch_related('entry_commission_rules')

    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if q:
        entries = entries.filter(
            Q(client__name__icontains=q) |
            Q(client__company__icontains=q) |
            Q(sales_rep__first_name__icontains=q) |
            Q(sales_rep__last_name__icontains=q) |
            Q(sales_rep__username__icontains=q) |
            Q(client__assigned_accountant__first_name__icontains=q) |
            Q(client__assigned_accountant__last_name__icontains=q) |
            Q(client__assigned_review__first_name__icontains=q) |
            Q(client__assigned_review__last_name__icontains=q)
        )
    if date_from:
        entries = entries.filter(entry_date__gte=date_from)
    if date_to:
        entries = entries.filter(entry_date__lte=date_to)

    totals = entries.aggregate(
        total_amount=Sum('amount'),
        total_commission=Sum('commission_amount'),
        total_accountant_commission=Sum('accountant_commission_amount'),
        total_review_commission=Sum('review_commission_amount'),
    )
    sales = totals['total_commission'] or 0
    accountant = totals['total_accountant_commission'] or 0
    review = totals['total_review_commission'] or 0
    totals['total_all_commissions'] = sales + accountant + review

    entries_list = list(entries)
    for entry in entries_list:
        rules = {r.department: r for r in entry.entry_commission_rules.all()}
        entry.sales_rule = rules.get('sales')
        entry.accountant_rule = rules.get('accountant')
        entry.review_rule = rules.get('review')

    return render(request, 'commissions/detail.html', {
        'sheet': sheet,
        'entries': entries_list,
        'all_entries': sheet.entries.only('pk'),
        'totals': totals,
        'q': q,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
@admin_required
def commission_save_entries(request, pk):
    sheet = get_object_or_404(CommissionSheet, pk=pk, tenant=request.user.tenant)
    if request.method == 'POST':
        entries = sheet.entries.prefetch_related('entry_commission_rules').all()
        for entry in entries:
            key = f'amount_{entry.pk}'
            if key not in request.POST:
                continue
            try:
                amount = Decimal(request.POST[key] or '0')
                if entry.amount != amount:
                    entry.amount = amount
                    entry.recalculate_commissions()
                    entry.save()
            except Exception:
                pass
        messages.success(request, 'تم حفظ جميع التعديلات')
    return redirect('commission_detail', pk=pk)


@login_required
@admin_required
def commission_toggle(request, pk):
    from django.db import transaction
    from django.db.models import Case, When, Value, BooleanField
    entry = get_object_or_404(CommissionEntry, pk=pk, sheet__tenant=request.user.tenant)
    with transaction.atomic():
        CommissionEntry.objects.filter(pk=pk).update(
            is_confirmed=Case(
                When(is_confirmed=True, then=Value(False)),
                default=Value(True),
                output_field=BooleanField(),
            )
        )
    return redirect('commission_detail', pk=entry.sheet.pk)


@login_required
def commission_delete_sheet(request, pk):
    if not request.user.is_admin:
        messages.error(request, 'هذه العملية للأدمن فقط')
        return redirect('commissions_list')
    sheet = get_object_or_404(CommissionSheet, pk=pk, tenant=request.user.tenant)
    from audit_log.utils import log_action
    from audit_log.models import AuditLog
    log_action(request, AuditLog.ACTION_DELETE, obj=sheet)
    sheet.delete()
    messages.success(request, 'تم حذف الشيت بنجاح')
    return redirect('commissions_list')


@login_required
@admin_required
def commission_refresh_sheet(request, pk):
    from clients.models import Client
    sheet = get_object_or_404(CommissionSheet, pk=pk, tenant=request.user.tenant)
    existing_ids = set(sheet.entries.values_list('client_id', flat=True))
    new_clients = Client.objects.filter(
        tenant=request.user.tenant,
        is_commissionable=True,
        is_active=True,
    ).exclude(id__in=existing_ids).select_related('assigned_sales')

    entries = [
        CommissionEntry(
            sheet=sheet,
            client=c,
            sales_rep=c.assigned_sales or request.user,
            amount=0,
        )
        for c in new_clients
    ]
    if entries:
        CommissionEntry.objects.bulk_create(entries)
        messages.success(request, f'تم إضافة {len(entries)} عميل جديد للشيت')
    else:
        messages.info(request, 'الشيت محدّث — لا يوجد عملاء جدد')
    return redirect('commission_detail', pk=pk)


@login_required
@admin_required
def commission_delete_entry(request, pk):
    entry = get_object_or_404(CommissionEntry, pk=pk, sheet__tenant=request.user.tenant)
    sheet_pk = entry.sheet.pk
    entry.delete()
    messages.success(request, 'تم حذف السطر')
    return redirect('commission_detail', pk=sheet_pk)


@login_required
@admin_required
def entry_amount_save(request, entry_pk):
    entry = get_object_or_404(CommissionEntry, pk=entry_pk, sheet__tenant=request.user.tenant)
    if request.method == 'POST':
        try:
            amount = Decimal(request.POST.get('amount', '0') or '0')
            if entry.amount != amount:
                old_amount = str(entry.amount)
                entry.amount = amount
                entry.recalculate_commissions()
                entry.save()
                from audit_log.utils import log_action
                from audit_log.models import AuditLog
                log_action(request, AuditLog.ACTION_UPDATE, obj=entry,
                           changes={'amount': {'من': old_amount, 'إلى': str(amount)}})
        except Exception:
            pass
    return redirect('commission_detail', pk=entry.sheet.pk)


@login_required
@admin_required
def entry_commission_save(request, entry_pk):
    entry = get_object_or_404(CommissionEntry, pk=entry_pk, sheet__tenant=request.user.tenant)
    department = request.POST.get('department')
    if request.method == 'POST' and department:
        EntryCommissionRule.objects.update_or_create(
            entry=entry,
            department=department,
            defaults={
                'commission_type': request.POST.get('commission_type', 'fixed'),
                'value': request.POST.get('value') or 0,
                'max_amount': request.POST.get('max_amount') or None,
                'reference_note': request.POST.get('reference_note', ''),
                'notes': request.POST.get('notes', ''),
            }
        )
        entry.recalculate_commissions()
        entry.save()
        from audit_log.utils import log_action
        from audit_log.models import AuditLog
        log_action(request, AuditLog.ACTION_UPDATE, obj=entry,
                   model_name='EntryCommissionRule',
                   object_repr=f'{entry} - {department}',
                   object_id=str(entry.pk))
        messages.success(request, 'تم حفظ قاعدة العمولة')
    return redirect('commission_detail', pk=entry.sheet.pk)


@login_required
@admin_required
def entry_commission_delete(request, entry_pk, department):
    entry = get_object_or_404(CommissionEntry, pk=entry_pk, sheet__tenant=request.user.tenant)
    EntryCommissionRule.objects.filter(entry=entry, department=department).delete()
    entry.recalculate_commissions()
    entry.save()
    from audit_log.utils import log_action
    from audit_log.models import AuditLog
    log_action(request, AuditLog.ACTION_DELETE,
               model_name='EntryCommissionRule',
               object_repr=f'{entry} - {department}',
               object_id=str(entry.pk))
    messages.success(request, 'تم حذف قاعدة العمولة')
    return redirect('commission_detail', pk=entry.sheet.pk)


# ═══════════════════════════════════════════
# تصدير شيت العمولات
# ═══════════════════════════════════════════

def _get_sheet_entries(sheet):
    return sheet.entries.select_related(
        'client', 'client__assigned_accountant', 'client__assigned_review', 'sales_rep'
    ).prefetch_related('entry_commission_rules').order_by('client__name')


def _apply_sheet_filters(entries, request):
    """Apply q / date_from / date_to filters from GET params."""
    q         = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if q:
        entries = entries.filter(
            Q(client__name__icontains=q) |
            Q(client__company__icontains=q) |
            Q(sales_rep__first_name__icontains=q) |
            Q(sales_rep__last_name__icontains=q) |
            Q(sales_rep__username__icontains=q) |
            Q(client__assigned_accountant__first_name__icontains=q) |
            Q(client__assigned_accountant__last_name__icontains=q) |
            Q(client__assigned_review__first_name__icontains=q) |
            Q(client__assigned_review__last_name__icontains=q)
        )
    if date_from:
        entries = entries.filter(entry_date__gte=date_from)
    if date_to:
        entries = entries.filter(entry_date__lte=date_to)
    return entries


@login_required
@admin_required
def export_sheet_excel(request, pk):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    sheet = get_object_or_404(CommissionSheet, pk=pk, tenant=request.user.tenant)
    entries = _apply_sheet_filters(_get_sheet_entries(sheet), request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet.name[:31]
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill('solid', fgColor='1A3A5C')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['العميل', 'الشركة', 'المندوب', 'المحاسب', 'المراجع', 'المبلغ',
               'عمولة المندوب', 'عمولة المحاسب', 'عمولة المراجع', 'الإجمالي', 'تم التأكيد']
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = 18

    for entry in entries:
        total = entry.commission_amount + entry.accountant_commission_amount + entry.review_commission_amount
        ws.append([
            entry.client.name,
            entry.client.company or '',
            entry.sales_rep.get_full_name() or entry.sales_rep.username,
            entry.client.assigned_accountant.get_full_name() if entry.client.assigned_accountant else '—',
            entry.client.assigned_review.get_full_name() if entry.client.assigned_review else '—',
            entry.amount,
            entry.commission_amount,
            entry.accountant_commission_amount,
            entry.review_commission_amount,
            total,
            'نعم' if entry.is_confirmed else 'لا',
        ])

    # صف المجاميع
    last = ws.max_row
    total_row = ['الإجمالي', '', '', '', '',
                 sum(e.amount for e in entries),
                 sum(e.commission_amount for e in entries),
                 sum(e.accountant_commission_amount for e in entries),
                 sum(e.review_commission_amount for e in entries),
                 sum(e.commission_amount + e.accountant_commission_amount + e.review_commission_amount for e in entries),
                 '']
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E8F0F8')

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sheet_{pk}.xlsx"'
    return response


@login_required
@admin_required
def export_sheet_pdf(request, pk):
    from reports.views import _render_pdf

    sheet    = get_object_or_404(CommissionSheet, pk=pk, tenant=request.user.tenant)
    entries  = list(_apply_sheet_filters(_get_sheet_entries(sheet), request))
    username = request.user.get_full_name() or request.user.username

    headers = ['العميل', 'الشركة', 'المندوب', 'المحاسب', 'المراجع', 'المبلغ', 'ع.المندوب', 'ع.المحاسب', 'ع.المراجع', 'الإجمالي']
    rows = []
    for e in entries:
        total = e.commission_amount + e.accountant_commission_amount + e.review_commission_amount
        rows.append([
            e.client.name,
            e.client.company or '',
            e.sales_rep.get_full_name() or e.sales_rep.username,
            e.client.assigned_accountant.get_full_name() if e.client.assigned_accountant else '—',
            e.client.assigned_review.get_full_name() if e.client.assigned_review else '—',
            f"{e.amount:,.2f}",
            f"{e.commission_amount:,.2f}",
            f"{e.accountant_commission_amount:,.2f}",
            f"{e.review_commission_amount:,.2f}",
            f"{total:,.2f}",
        ])
    rows.append([
        'الإجمالي', '', '', '', '',
        f"{sum(e.amount for e in entries):,.2f}",
        f"{sum(e.commission_amount for e in entries):,.2f}",
        f"{sum(e.accountant_commission_amount for e in entries):,.2f}",
        f"{sum(e.review_commission_amount for e in entries):,.2f}",
        f"{sum(e.commission_amount + e.accountant_commission_amount + e.review_commission_amount for e in entries):,.2f}",
    ])
    return _render_pdf(sheet.name, headers, rows, f'sheet_{pk}.pdf', username, has_totals=True)
